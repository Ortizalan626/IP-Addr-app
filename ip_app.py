import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import subprocess
import re
from datetime import datetime
import tempfile
import os
import threading
import queue
import time

# ----------------------------- NET/VALIDATION -----------------------------
IP_PATTERN = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$")
CREATE_NO_WINDOW = 0x08000000

BATCH_SIZE = 25

CHECK_OFF = "☐"
CHECK_ON = "☑"


def is_valid_ip(ip: str) -> bool:
    ip = ip.strip()
    if not IP_PATTERN.match(ip):
        return False
    try:
        return all(0 <= int(x) <= 255 for x in ip.split("."))
    except:
        return False


def run_hidden(cmd_list):
    return subprocess.run(
        cmd_list,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        creationflags=CREATE_NO_WINDOW
    )


def popen_hidden(cmd_list):
    return subprocess.Popen(
        cmd_list,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        creationflags=CREATE_NO_WINDOW
    )


def is_admin():
    try:
        r = run_hidden(["netsh", "interface", "ipv4", "show", "interfaces"])
        return r.returncode == 0
    except:
        return False


def get_nics():
    try:
        r = run_hidden(["netsh", "interface", "show", "interface"])
        nics = []
        for line in r.stdout.splitlines():
            parts = line.split()
            if len(parts) >= 4 and parts[0] in ("Enabled", "Disabled"):
                nics.append(" ".join(parts[3:]))
        return nics
    except:
        return ["Ethernet"]


def _get_assigned_ips_powershell(nic):
    try:
        safe_nic = nic.replace("'", "''")
        ps_cmd = (
            f"try {{ "
            f"Get-NetIPAddress -InterfaceAlias '{safe_nic}' -AddressFamily IPv4 -ErrorAction Stop "
            f"| Select-Object -ExpandProperty IPAddress "
            f"}} catch {{ }}"
        )
        r = run_hidden(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_cmd])
        if r.returncode != 0:
            return None

        ips = []
        for line in (r.stdout or "").splitlines():
            ip = line.strip()
            if not ip:
                continue
            if not is_valid_ip(ip):
                continue
            if ip.startswith("169.254.") or ip in ("0.0.0.0",):
                continue
            ips.append(ip)
        return ips
    except:
        return None


def _get_assigned_ips_netsh_parse(nic):
    try:
        r = run_hidden(["netsh", "interface", "ipv4", "show", "addresses", f'name="{nic}"'])
        if r.returncode != 0:
            r = run_hidden(["netsh", "interface", "ipv4", "show", "addresses", f"name={nic}"])
        if r.returncode != 0:
            return []

        ips = []
        for line in (r.stdout or "").splitlines():
            line = line.strip()
            if line.lower().startswith("ip address"):
                raw = line.split(":")[-1].strip()
                raw = raw.split("(")[0].strip()
                if is_valid_ip(raw) and not raw.startswith("169.254.") and raw != "0.0.0.0":
                    ips.append(raw)
        return ips
    except:
        return []


def get_assigned_ips(nic, prefer: str = "powershell"):
    """
    Return IPv4 addresses assigned to the given NIC (InterfaceAlias).

    prefer:
      - "powershell" (default): PowerShell first, then netsh fallback
      - "netsh": netsh first (often faster on locked-down/corporate PCs), then PowerShell fallback
    """
    prefer = (prefer or "powershell").lower().strip()

    if prefer == "netsh":
        ips = _get_assigned_ips_netsh_parse(nic)
        if ips:
            return ips
        ips2 = _get_assigned_ips_powershell(nic)
        return ips2 if ips2 is not None else ips

    # default: PowerShell first (more direct API), then netsh fallback
    ips = _get_assigned_ips_powershell(nic)
    if ips is not None:
        return ips
    return _get_assigned_ips_netsh_parse(nic)


# ----------------------------- TEXT WIDGET -----------------------------
class SafeScrolledText(tk.Frame):
    def __init__(self, master, width=40, height=10):
        super().__init__(master)
        self.text = tk.Text(self, width=width, height=height, wrap="none")
        self.scroll = tk.Scrollbar(self, command=self.text.yview)
        self.text.configure(yscrollcommand=self.scroll.set)

        self.text.pack(side="left", fill="both", expand=True)
        self.scroll.pack(side="right", fill="y")

        self._readonly = False

    def get_lines(self):
        raw = self.text.get("1.0", "end").strip()
        out = []
        for ln in raw.splitlines():
            ln = ln.strip()
            if not ln:
                continue
            if ln.startswith(CHECK_OFF) or ln.startswith(CHECK_ON):
                ln = ln[1:].strip()
            out.append(ln)
        return out

    def set_text(self, lines):
        self.text.configure(state="normal")
        self.text.delete("1.0", "end")
        if lines:
            self.text.insert("end", "\n".join(lines) + "\n")
        self.text.see("end")
        self.text.configure(state="disabled" if self._readonly else "normal")

    def append_line(self, line: str):
        self.text.configure(state="normal")
        self.text.insert("end", line + "\n")
        self.text.see("end")
        self.text.configure(state="disabled" if self._readonly else "normal")

    def clear(self):
        self.text.configure(state="normal")
        self.text.delete("1.0", "end")
        self.text.see("end")
        self.text.configure(state="disabled" if self._readonly else "normal")

    def set_readonly(self, ro: bool):
        self._readonly = ro
        self.text.configure(state="disabled" if ro else "normal")


# ----------------------------- MAIN APP -----------------------------
class IPUtilityApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("IP Utility")
        self.root.geometry("1000x780")
        self.logs = []
        self._op = None
        self.assigned_checks = {}

        # ------------------ TOP BAR ------------------
        self.top = tk.Frame(self.root)
        self.top.pack(fill="x", padx=10, pady=10)

        tk.Label(self.top, text="Network Adapter:").pack(side="left", padx=(0, 6))

        self.nic_var = tk.StringVar()
        self.nic_combo = ttk.Combobox(self.top, textvariable=self.nic_var, width=35, state="readonly")
        self.nic_combo.pack(side="left", padx=(0, 20))

        tk.Label(self.top, text="Subnet Mask:").pack(side="left", padx=(0, 6))

        self.mask_entry = tk.Entry(self.top, width=20)
        self.mask_entry.insert(0, "255.255.254.0")
        self.mask_entry.pack(side="left", padx=(0, 20))

        self.theme_var = tk.StringVar(value="Dark Mode")
        self.theme_combo = ttk.Combobox(
            self.top, textvariable=self.theme_var,
            values=["Dark Mode", "Light Mode"],
            width=20, state="readonly"
        )
        self.theme_combo.pack(side="right")

        # ------------------ MAIN COLUMNS ------------------
        self.columns = tk.Frame(self.root)
        self.columns.pack(fill="x", padx=10)

        # LEFT SIDE
        self.left = tk.Frame(self.columns)
        self.left.pack(side="left", padx=20)

        self.left_label = tk.Label(self.left, text="Paste IPs Here")
        self.left_label.pack(anchor="w")

        self.input_box = SafeScrolledText(self.left, width=48, height=18)
        self.input_box.pack()
        self.input_box.set_readonly(False)

        self.left_btns = tk.Frame(self.left)
        self.left_btns.pack(fill="x", pady=10)

        self.add_btn = tk.Button(self.left_btns, text="Add IPs to Adapter", command=self.add_ips)
        self.add_btn.pack(side="left", expand=True, fill="x")

        self.clear_btn = tk.Button(self.left_btns, text="Clear List", command=self.input_box.clear)
        self.clear_btn.pack(side="left", expand=False, fill="x", padx=2)

        self.excel_btn = tk.Button(self.left, text="Load from Excel", command=self.load_excel)
        self.excel_btn.pack(fill="x")

        # RIGHT SIDE
        self.right = tk.Frame(self.columns)
        self.right.pack(side="left", padx=20)

        self.assigned_label = tk.Label(self.right, text="IPs Currently Assigned to Adapter:")
        self.assigned_label.pack(anchor="w")

        self.assigned_box = SafeScrolledText(self.right, width=48, height=18)
        self.assigned_box.pack()
        self.assigned_box.set_readonly(True)

        # Remove row (horizontal buttons)
        self.remove_row = tk.Frame(self.right)
        self.remove_row.pack(fill="x", pady=(10, 5))

        self.remove_btn = tk.Button(self.remove_row, text="Remove ALL Assigned IPs from Adapter", command=self.remove_all)
        self.remove_btn.pack(side="left", expand=True, fill="x")

        self.remove_selected_btn = tk.Button(self.remove_row, text="Remove Selected", command=self.remove_selected)
        self.remove_selected_btn.pack(side="left", expand=False, fill="x", padx=(6, 0))

        self.refresh_btn = tk.Button(self.right, text="Refresh Assigned List", command=self.refresh_assigned)
        self.refresh_btn.pack(fill="x")

        # ------------------ LOG AREA ------------------
        self.log_area = tk.Frame(self.root)
        self.log_area.pack(fill="both", expand=True, padx=10, pady=10)

        self.log_header = tk.Frame(self.log_area)
        self.log_header.pack(fill="x", anchor="w")

        self.log_label = tk.Label(self.log_header, text="Log:")
        self.log_label.pack(side="left", anchor="w")

        self.progress = ttk.Progressbar(self.log_header, orient="horizontal", mode="determinate", length=260)
        self.progress.pack(side="right", padx=(10, 0), pady=(0, 2))
        self.progress["maximum"] = 1
        self.progress["value"] = 0

        self.log_box = SafeScrolledText(self.log_area, width=120, height=10)
        self.log_box.pack(fill="both", expand=True)
        self.log_box.set_readonly(True)

        self.clear_log_btn = tk.Button(self.log_area, text="Clear Log", command=self.clear_log)
        self.clear_log_btn.pack(anchor="e", pady=(5, 0))

        # ------------------ EVENT BINDINGS ------------------
        self.nic_combo.bind("<<ComboboxSelected>>", self.on_nic_change)
        self.theme_combo.bind("<<ComboboxSelected>>", lambda e: self.apply_theme())
        self.assigned_box.text.bind("<Button-1>", self._on_assigned_click)

        # ------------------ INIT (FAST STARTUP) ------------------
        self.apply_theme()

        # IMPORTANT: do slow netsh / powershell calls off the UI thread
        self.nic_combo["values"] = ["(Loading adapters...)"]
        try:
            self.nic_combo.current(0)
        except:
            pass

        self._startup_q = queue.Queue()
        self._refresh_q = queue.Queue()
        self._startup_thread = None
        self._refresh_thread = None
        self._last_log_flush = 0.0

        self.root.after(10, self._startup_async)

    # ------------------ STARTUP LOAD (DEFERRED) ------------------
    def _startup_load(self):
        # populate NICs (netsh can be slow)
        self.nic_combo["values"] = get_nics()
        if self.nic_combo["values"]:
            self.nic_combo.current(0)
            self.nic_var.set(self.nic_combo.get())

        # refresh assigned list (PowerShell can be slow)
        self.refresh_assigned()

    # ------------------ ASYNC STARTUP / REFRESH ------------------
    def _startup_async(self):
        # avoid starting multiple startup threads
        if self._startup_thread is not None and self._startup_thread.is_alive():
            return

        self._startup_thread = threading.Thread(target=self._startup_worker, daemon=True)
        self._startup_thread.start()
        self.root.after(50, self._startup_poll)

    def _startup_worker(self):
        # run slow OS calls off the UI thread
        nics = get_nics()
        first = nics[0] if nics else ""
        ips = get_assigned_ips(first, prefer="netsh") if first else []
        self._startup_q.put(("startup", nics, first, ips))

    def _startup_poll(self):
        try:
            _, nics, first, ips = self._startup_q.get_nowait()
        except queue.Empty:
            if self._startup_thread is not None and self._startup_thread.is_alive():
                self.root.after(50, self._startup_poll)
            return

        # update UI
        self.nic_combo["values"] = nics if nics else ["(No adapters found)"]
        if nics:
            self.nic_var.set(first)
            try:
                self.nic_combo.current(0)
            except:
                pass

        old = dict(self.assigned_checks)
        self.assigned_checks = {ip: bool(old.get(ip, False)) for ip in ips}
        self._render_assigned_with_checks(ips)

        if first:
            self.assigned_label.config(text=f"IPs Currently Assigned to Adapter: {first}")
        else:
            self.assigned_label.config(text="IPs Currently Assigned to Adapter:")

    def refresh_assigned(self):
        # async refresh to keep UI responsive on slower/locked-down machines
        nic = self.nic_var.get().strip()
        if not nic:
            self.assigned_box.set_text([])
            return

        # don't refresh while add/remove netsh batches are in flight
        if self._op is not None:
            return

        if self._refresh_thread is not None and self._refresh_thread.is_alive():
            return

        self._set_busy(True)
        self._refresh_thread = threading.Thread(target=self._refresh_worker, args=(nic,), daemon=True)
        self._refresh_thread.start()
        self.root.after(50, self._refresh_poll)

    def _refresh_worker(self, nic: str):
        ips = get_assigned_ips(nic, prefer="netsh")
        self._refresh_q.put(("refresh", nic, ips))

    def _refresh_poll(self):
        try:
            _, nic, ips = self._refresh_q.get_nowait()
        except queue.Empty:
            if self._refresh_thread is not None and self._refresh_thread.is_alive():
                self.root.after(50, self._refresh_poll)
            else:
                self._set_busy(False)
            return

        old = dict(self.assigned_checks)
        self.assigned_checks = {ip: bool(old.get(ip, False)) for ip in ips}
        self._render_assigned_with_checks(ips)
        self.assigned_label.config(text=f"IPs Currently Assigned to Adapter: {nic}")
        self._set_busy(False)


    # --------------------------------------------------
    # Assigned list w/ checkboxes
    # --------------------------------------------------
    def _render_assigned_with_checks(self, ips):
        lines = []
        for ip in ips:
            checked = bool(self.assigned_checks.get(ip, False))
            lines.append(f"{CHECK_ON if checked else CHECK_OFF} {ip}")
        self.assigned_box.set_text(lines)

    def _on_assigned_click(self, event):
        t = self.assigned_box.text
        try:
            idx = t.index(f"@{event.x},{event.y}")
            line_s, col_s = idx.split(".")
            line = int(line_s)
            col = int(col_s)
        except:
            return

        # only toggle if click in checkbox area
        if col > 1:
            return

        line_text = t.get(f"{line}.0", f"{line}.end").strip()
        if not line_text:
            return

        parts = line_text.split(maxsplit=1)
        if len(parts) != 2:
            return

        ip = parts[1].strip()
        if not is_valid_ip(ip):
            return

        new_state = not bool(self.assigned_checks.get(ip, False))
        self.assigned_checks[ip] = new_state

        t.configure(state="normal")
        try:
            t.delete(f"{line}.0", f"{line}.1")
            t.insert(f"{line}.0", CHECK_ON if new_state else CHECK_OFF)
        finally:
            t.configure(state="disabled")

        return "break"

    def _get_selected_assigned_ips(self):
        return [ip for ip, checked in self.assigned_checks.items() if checked]

    # --------------------------------------------------
    # THEME ENGINE (Text-safe)
    # --------------------------------------------------
    def apply_theme(self):
        if self.theme_var.get() == "Light Mode":
            self.apply_light()
        else:
            self.apply_dark()

    def _apply_text_colors(self, widget, bg, fg, insertbg):
        txt = widget.text
        old_state = txt.cget("state")
        txt.configure(state="normal")
        txt.configure(bg=bg, fg=fg, insertbackground=insertbg)
        txt.configure(state=old_state)

    def apply_dark(self):
        bg = "#121212"
        fg = "#ffffff"
        boxbg = "#1e1e1e"

        for w in [self.root, self.top, self.columns, self.left, self.right, self.log_area, self.log_header, self.remove_row]:
            w.configure(bg=bg)

        for lbl in [self.left_label, self.assigned_label, self.log_label]:
            lbl.configure(bg=bg, fg=fg)

        self.mask_entry.configure(bg=boxbg, fg=fg, insertbackground="white")

        for s in [self.input_box, self.assigned_box, self.log_box]:
            self._apply_text_colors(s, boxbg, fg, "white")
            s.scroll.configure(bg=bg)

        self.left_btns.configure(bg=bg)

        self.add_btn.configure(bg="#005a9e", fg="white")
        self.clear_btn.configure(bg="#005a9e", fg="white")
        self.excel_btn.configure(bg="#3a3a3a", fg="white")
        self.remove_btn.configure(bg="#b83838", fg="white")
        self.remove_selected_btn.configure(bg="#b83838", fg="white")
        self.refresh_btn.configure(bg="#3a3a3a", fg="white")

    def apply_light(self):
        bg = "#ffffff"
        fg = "#000000"
        boxbg = "#ffffff"

        for w in [self.root, self.top, self.columns, self.left, self.right, self.log_area, self.log_header, self.remove_row]:
            w.configure(bg=bg)

        for lbl in [self.left_label, self.assigned_label, self.log_label]:
            lbl.configure(bg=bg, fg=fg)

        self.mask_entry.configure(bg=boxbg, fg=fg, insertbackground="black")

        for s in [self.input_box, self.assigned_box, self.log_box]:
            self._apply_text_colors(s, boxbg, fg, "black")
            s.scroll.configure(bg="#efefef")

        self.left_btns.configure(bg=bg)

        self.add_btn.configure(bg="#d4e8ff", fg="black")
        self.clear_btn.configure(bg="#d4e8ff", fg="black")
        self.excel_btn.configure(bg="#d0d0d0", fg="black")
        self.remove_btn.configure(bg="#f2b3b3", fg="black")
        self.remove_selected_btn.configure(bg="#f2b3b3", fg="black")
        self.refresh_btn.configure(bg="#d0d0d0", fg="black")

    # --------------------------------------------------
    # LOGGING
    # --------------------------------------------------
    def log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        entry = f"[{ts}] {msg}"
        self.logs.append(entry)
        if len(self.logs) > 500:
            self.logs = self.logs[-500:]
        self.log_box.append_line(entry)

        # Throttle GUI flushes (massive speedup when adding/removing many IPs)
        now = time.time()
        if now - getattr(self, "_last_log_flush", 0.0) >= 0.05:
            self._last_log_flush = now
            self.root.update_idletasks()

    def clear_log(self):
        self.logs = []
        self.log_box.set_readonly(False)
        self.log_box.clear()
        self.log_box.set_readonly(True)

    # --------------------------------------------------
    # Progress / Busy
    # --------------------------------------------------
    def _progress_reset(self):
        self.progress["maximum"] = 1
        self.progress["value"] = 0
        self.root.update_idletasks()

    def _progress_start(self, total: int, initial: int = 0):
        self.progress["maximum"] = max(1, int(total))
        self.progress["value"] = int(initial)
        self.root.update_idletasks()

    def _progress_add(self, delta: int):
        self.progress["value"] = int(self.progress["value"] + int(delta))
        self.root.update_idletasks()

    def _set_busy(self, busy: bool):
        state_btn = "disabled" if busy else "normal"
        self.add_btn.configure(state=state_btn)
        self.remove_btn.configure(state=state_btn)
        self.remove_selected_btn.configure(state=state_btn)
        self.refresh_btn.configure(state=state_btn)
        self.excel_btn.configure(state=state_btn)
        self.clear_btn.configure(state=state_btn)
        self.nic_combo.configure(state="disabled" if busy else "readonly")
        self.theme_combo.configure(state="disabled" if busy else "readonly")
        self.mask_entry.configure(state="disabled" if busy else "normal")
        self.root.update_idletasks()

    # --------------------------------------------------
    # NIC CHANGE / REFRESH
    # --------------------------------------------------
    def on_nic_change(self, event=None):
        nic = self.nic_var.get().strip()
        self.assigned_label.config(text=f"IPs Currently Assigned to Adapter: {nic}")
        self.refresh_assigned()


    # --------------------------------------------------
    # Batch netsh helpers
    # --------------------------------------------------
    def _write_netsh_script(self, lines):
        tf = tempfile.NamedTemporaryFile(delete=False, mode="w", encoding="utf-8", suffix=".txt")
        try:
            for ln in lines:
                tf.write(ln.rstrip() + "\n")
            tf.flush()
            return tf.name
        finally:
            tf.close()

    def _start_batch_process(self, script_path):
        return popen_hidden(["netsh", "-f", script_path])

    def _cleanup_script(self, path):
        try:
            os.remove(path)
        except:
            pass

    # --------------------------------------------------
    # ADD IPS (batched)
    # --------------------------------------------------
    def add_ips(self):
        if self._op is not None:
            return

        if not is_admin():
            messagebox.showerror("Error", "Run this program as Administrator.")
            return

        nic = self.nic_var.get().strip()
        mask = self.mask_entry.get().strip()

        raw_ips = [ip for ip in self.input_box.get_lines() if is_valid_ip(ip)]
        if not raw_ips:
            messagebox.showerror("Error", "No valid IPs found.")
            return

        existing = set(get_assigned_ips(nic, prefer="netsh"))

        pending = []
        skipped = 0
        for ip in raw_ips:
            if ip in existing:
                self.log(f"SKIP (already assigned): {ip}")
                skipped += 1
            else:
                pending.append(ip)

        total = len(raw_ips)

        self._set_busy(True)
        self._progress_start(total, initial=skipped)

        if not pending:
            self.log(f"Finished. Added 0, skipped {skipped}.")
            self.refresh_assigned()
            self.input_box.clear()
            self._progress_reset()
            self._set_busy(False)
            return

        self._op = {
            "type": "add",
            "nic": nic,
            "mask": mask,
            "pending": pending,
            "index": 0,
            "skipped": skipped,
            "added": 0,
            "errors": 0,
            "proc": None,
            "script": None,
            "batch_ips": None,
        }

        self._add_run_next_batch()

    def _add_run_next_batch(self):
        op = self._op
        if not op or op["type"] != "add":
            return

        if op["proc"] is not None:
            self._add_poll_proc()
            return

        pending = op["pending"]
        idx = op["index"]
        if idx >= len(pending):
            self.log(f"Finished. Added {op['added']}, skipped {op['skipped']}, errors {op['errors']}.")
            self._op = None
            self.refresh_assigned()
            self.input_box.clear()
            self._progress_reset()
            self._set_busy(False)
            return

        batch = pending[idx: idx + BATCH_SIZE]
        op["batch_ips"] = batch
        op["index"] = idx + len(batch)

        script_lines = []
        for ip in batch:
            self.log(f"ADDING: {ip}")
            script_lines.append(
                f'interface ipv4 add address name="{op["nic"]}" address={ip} mask={op["mask"]} store=persistent'
            )

        script_path = self._write_netsh_script(script_lines)
        op["script"] = script_path
        op["proc"] = self._start_batch_process(script_path)

        self.root.after(50, self._add_poll_proc)

    def _add_poll_proc(self):
        op = self._op
        if not op or op["type"] != "add" or op["proc"] is None:
            return

        p = op["proc"]
        rc = p.poll()
        if rc is None:
            self.root.after(75, self._add_poll_proc)
            return

        out, err = p.communicate()
        self._cleanup_script(op["script"])

        op["proc"] = None
        op["script"] = None

        current = set(get_assigned_ips(op["nic"], prefer="netsh"))

        for ip in op["batch_ips"]:
            if ip in current:
                self.log(f"SUCCESS: {ip}")
                op["added"] += 1
            else:
                details = (err or out or "").strip()
                if not details:
                    details = f"(netsh completed rc={rc}, but IP not observed after batch)"
                self.log(f"ERROR adding {ip}: {details}")
                op["errors"] += 1

        self._progress_add(len(op["batch_ips"]))
        op["batch_ips"] = None

        self.root.after(1, self._add_run_next_batch)

    # --------------------------------------------------
    # REMOVE helpers
    # --------------------------------------------------
    def _begin_remove(self, ips, mode_label: str):
        if self._op is not None:
            return False

        if not is_admin():
            messagebox.showerror("Error", "Run this program as Administrator.")
            return False

        nic = self.nic_var.get().strip()

        if not ips:
            self.log("No IPs to remove.")
            self._progress_reset()
            return False

        self.log(f"Removing {len(ips)} IPs ({mode_label})...")

        self._set_busy(True)
        self._progress_start(len(ips), initial=0)

        self._op = {
            "type": "remove",
            "mode": mode_label,
            "nic": nic,
            "ips": ips,
            "index": 0,
            "removed": 0,
            "errors": 0,
            "proc": None,
            "script": None,
            "batch_ips": None,
        }

        self._remove_run_next_batch()
        return True

    def remove_all(self):
        nic = self.nic_var.get().strip()
        ips = get_assigned_ips(nic, prefer="netsh")
        self._begin_remove(ips, "all")

    def remove_selected(self):
        selected = self._get_selected_assigned_ips()
        if not selected:
            messagebox.showinfo("Remove Selected", "No IPs are selected.")
            return

        nic = self.nic_var.get().strip()
        current = set(get_assigned_ips(nic, prefer="netsh"))
        selected = [ip for ip in selected if ip in current]

        if not selected:
            messagebox.showinfo("Remove Selected", "Selected IPs are not currently assigned (refresh list).")
            self.refresh_assigned()
            return

        self._begin_remove(selected, "selected")

    def _remove_run_next_batch(self):
        op = self._op
        if not op or op["type"] != "remove":
            return

        if op["proc"] is not None:
            self._remove_poll_proc()
            return

        ips = op["ips"]
        idx = op["index"]
        if idx >= len(ips):
            if op.get("mode") == "selected":
                self.log(f"Finished removing selected: {op['removed']} removed. Errors {op['errors']}.")
            else:
                self.log(f"Finished removing {op['removed']} IPs. Errors {op['errors']}.")
            self._op = None
            self.refresh_assigned()
            self._progress_reset()
            self._set_busy(False)
            return

        batch = ips[idx: idx + BATCH_SIZE]
        op["batch_ips"] = batch
        op["index"] = idx + len(batch)

        script_lines = []
        for ip in batch:
            self.log(f"DELETING: {ip}")
            script_lines.append(f'interface ipv4 delete address name="{op["nic"]}" address={ip}')

        script_path = self._write_netsh_script(script_lines)
        op["script"] = script_path
        op["proc"] = self._start_batch_process(script_path)

        self.root.after(50, self._remove_poll_proc)

    def _remove_poll_proc(self):
        op = self._op
        if not op or op["type"] != "remove" or op["proc"] is None:
            return

        p = op["proc"]
        rc = p.poll()
        if rc is None:
            self.root.after(75, self._remove_poll_proc)
            return

        out, err = p.communicate()
        self._cleanup_script(op["script"])

        op["proc"] = None
        op["script"] = None

        current = set(get_assigned_ips(op["nic"], prefer="netsh"))

        for ip in op["batch_ips"]:
            if ip not in current:
                self.log(f"SUCCESS removing {ip}")
                op["removed"] += 1
                if ip in self.assigned_checks:
                    self.assigned_checks[ip] = False
            else:
                details = (err or out or "").strip()
                if not details:
                    details = f"(netsh completed rc={rc}, but IP still observed after batch)"
                self.log(f"ERROR removing {ip}: {details}")
                op["errors"] += 1

        self._progress_add(len(op["batch_ips"]))
        op["batch_ips"] = None

        self.root.after(1, self._remove_run_next_batch)

    # --------------------------------------------------
    # EXCEL IMPORT (lazy import openpyxl)
    # --------------------------------------------------
    def load_excel(self):
        file = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        if not file:
            return

        try:
            import openpyxl  # lazy-load to speed up app startup
        except:
            messagebox.showerror("Error", "openpyxl is not installed. Run: pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except:
            messagebox.showerror("Error", "Could not open file.")
            return

        sheet = None
        for name in wb.sheetnames:
            if name.lower() == "sources":
                sheet = wb[name]
                break

        if sheet is None:
            messagebox.showerror("Error", "Sheet 'Sources' not found.")
            return

        header_row = None
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if not row:
                continue
            lower = [str(c).strip().lower() if c else "" for c in row]
            if "routing add" in lower:
                header_row = lower
                break

        if header_row is None:
            messagebox.showerror("Error", "Column 'ROUTING ADD' not found.")
            return

        col_index = header_row.index("routing add") + 1

        ips = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            val = row[col_index - 1]
            if val and isinstance(val, str) and is_valid_ip(val):
                ips.append(val)

        self.input_box.set_text(ips)
        self.log(f"Loaded {len(ips)} IPs from Excel.")

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    IPUtilityApp().run()
